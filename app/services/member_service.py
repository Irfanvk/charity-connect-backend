import csv
import io
import re
import secrets
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session

from app.models import Member, User, Challan, BulkChallanGroup, Campaign
from app.schemas import MemberUpdate, MemberCreate, MemberImportSummary
from fastapi import HTTPException, status
from app.utils import hash_password, generate_member_code


class MemberService:
    """Member management service."""

    @staticmethod
    def _normalize_member_code(code: Optional[str]) -> Optional[str]:
        if not code:
            return None
        return str(code).strip().upper()

    @staticmethod
    def normalize_member_code(code: Optional[str]) -> Optional[str]:
        return MemberService._normalize_member_code(code)

    @staticmethod
    def _normalize_contact(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        cleaned = str(value).strip()
        return cleaned or None

    @staticmethod
    def normalize_contact(value: Optional[str]) -> Optional[str]:
        return MemberService._normalize_contact(value)

    @staticmethod
    def _safe_username_seed(name: Optional[str], fallback: str = "member") -> str:
        base = re.sub(r"[^a-zA-Z0-9]+", "_", (name or "").strip().lower()).strip("_")
        if not base:
            base = fallback
        return base[:24]

    @staticmethod
    def _generate_unique_offline_username(db: Session, full_name: Optional[str], member_code: str) -> str:
        seed = MemberService._safe_username_seed(full_name, fallback="member")
        code_seed = re.sub(r"[^a-zA-Z0-9]", "", member_code.lower())[-6:] or "mem"

        for _ in range(20):
            suffix = secrets.token_hex(2)
            candidate = f"offline_{seed}_{code_seed}_{suffix}"[:255]
            exists = db.query(User).filter(User.username == candidate).first()
            if not exists:
                return candidate

        return f"offline_{secrets.token_hex(8)}"

    @staticmethod
    def _next_member_code(db: Session) -> str:
        last_member = db.query(Member).order_by(Member.id.desc()).first()
        last_code = last_member.member_code if last_member else None
        return generate_member_code(last_code)

    @staticmethod
    def _get_or_create_user_for_admin_member(
        db: Session,
        full_name: str,
        phone: Optional[str],
        email: Optional[str],
        member_code: str,
    ) -> tuple[User, bool]:
        """Return (user, created_new_user). Reuses existing user by contact when possible."""
        existing_user = None

        if email:
            existing_user = db.query(User).filter(User.email == email).first()
        if not existing_user and phone:
            existing_user = db.query(User).filter(User.phone == phone).first()

        if existing_user:
            return existing_user, False

        username = MemberService._generate_unique_offline_username(db, full_name, member_code)
        offline_secret = secrets.token_urlsafe(24)
        user = User(
            username=username,
            email=email,
            phone=phone,
            password_hash=hash_password(offline_secret),
            role="member",
            is_active=False,
        )
        db.add(user)
        db.flush()
        return user, True

    @staticmethod
    def _parse_datetime(value) -> Optional[datetime]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        raw = str(value).strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(raw)
        except ValueError:
            return None

    @staticmethod
    def _si_to_member_code(si_value) -> Optional[str]:
        if si_value in (None, ""):
            return None

        try:
            si_int = int(float(str(si_value).strip()))
            return f"MEM-{si_int:04d}"
        except (TypeError, ValueError):
            return None

    @staticmethod
    def si_to_member_code(si_value) -> Optional[str]:
        return MemberService._si_to_member_code(si_value)

    @staticmethod
    def _parse_month(value: Optional[str]) -> Optional[str]:
        raw = MemberService._normalize_contact(value)
        if not raw:
            return None

        # Canonical format used by challans: YYYY-MM
        if re.match(r"^\d{4}-\d{2}$", raw):
            return raw

        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).strftime("%Y-%m")
            except ValueError:
                continue

        return None

    @staticmethod
    def parse_month(value: Optional[str]) -> Optional[str]:
        return MemberService._parse_month(value)

    @staticmethod
    def _find_member_by_identifiers(
        db: Session,
        member_code: Optional[str],
        username: Optional[str],
        phone: Optional[str],
        email: Optional[str],
    ) -> Optional[Member]:
        if member_code:
            member = db.query(Member).filter(Member.member_code == member_code).first()
            if member:
                return member

        if username:
            user = db.query(User).filter(User.username == username).first()
            if user:
                member = db.query(Member).filter(Member.user_id == user.id).first()
                if member:
                    return member

        if email:
            user = db.query(User).filter(User.email == email).first()
            if user:
                member = db.query(Member).filter(Member.user_id == user.id).first()
                if member:
                    return member

        if phone:
            user = db.query(User).filter(User.phone == phone).first()
            if user:
                member = db.query(Member).filter(Member.user_id == user.id).first()
                if member:
                    return member

        return None

    @staticmethod
    def find_member_by_identifiers(
        db: Session,
        member_code: Optional[str],
        username: Optional[str],
        phone: Optional[str],
        email: Optional[str],
    ) -> Optional[Member]:
        return MemberService._find_member_by_identifiers(
            db=db,
            member_code=member_code,
            username=username,
            phone=phone,
            email=email,
        )

    @staticmethod
    def _resolve_campaign_for_import(
        db: Session,
        suggested_campaign_name: Optional[str],
        imported_by_user_id: Optional[int],
    ) -> Optional[int]:
        title = MemberService._normalize_contact(suggested_campaign_name)
        if not title:
            return None

        existing = db.query(Campaign).filter(Campaign.title == title).first()
        if existing:
            return existing.id

        created_by_user_id = imported_by_user_id
        if not created_by_user_id:
            admin_user = db.query(User).filter(User.role.in_(["superadmin", "admin"]))\
                .order_by(User.id.asc()).first()
            created_by_user_id = admin_user.id if admin_user else None

        if not created_by_user_id:
            return None

        now = datetime.utcnow()
        campaign = Campaign(
            title=title,
            description="Imported historical one-time/campaign donations",
            target_amount=0.0,
            start_date=now,
            end_date=now,
            is_active=False,
            created_by_admin_id=created_by_user_id,
        )
        db.add(campaign)
        db.flush()
        return campaign.id

    @staticmethod
    def resolve_campaign_for_import(
        db: Session,
        suggested_campaign_name: Optional[str],
        imported_by_user_id: Optional[int],
    ) -> Optional[int]:
        return MemberService._resolve_campaign_for_import(
            db=db,
            suggested_campaign_name=suggested_campaign_name,
            imported_by_user_id=imported_by_user_id,
        )

    @staticmethod
    def row_value(row: dict, keys: list[str]):
        return MemberService._row_value(row, keys)

    @staticmethod
    def normalized_row(row: dict) -> dict:
        return MemberService._normalized_row(row)

    @staticmethod
    def read_tabular_rows(file_bytes: bytes, filename: str) -> list[dict]:
        return MemberService._read_tabular_rows(file_bytes, filename)

    @staticmethod
    def get_member(db: Session, member_id: int):
        """Get member by ID."""
        member = db.query(Member).filter(Member.id == member_id).first()

        if not member:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Member not found",
            )

        return member

    @staticmethod
    def get_member_by_code(db: Session, member_code: str):
        """Get member by member code."""
        member = db.query(Member).filter(Member.member_code == member_code).first()

        if not member:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Member not found",
            )

        return member

    @staticmethod
    def get_all_members(
        db: Session,
        skip: int = 0,
        limit: int = 20,
        search: Optional[str] = None,
        sort_by: str = "full_name",
        sort_order: str = "asc",
    ):
        """Get members with search, sorting, and pagination."""
        from sqlalchemy import asc, desc

        # Always join User so we can filter/sort on User.username
        query = db.query(Member).join(User, Member.user_id == User.id)

        # Search against User.username (the real name column) and Member.member_code
        if search:
            query = query.filter(
                User.username.ilike(f"%{search}%") |
                Member.member_code.ilike(f"%{search}%")
            )

        # Sorting — use actual mapped columns, never @property descriptors
        if sort_by in ("name", "full_name"):
            column = User.username
        elif sort_by == "id":
            column = Member.member_code
        else:
            column = User.username

        order_fn = desc if sort_order == "desc" else asc
        query = query.order_by(order_fn(column))

        # Pagination
        return query.offset(skip).limit(limit).all()

    @staticmethod
    def create_member(db: Session, member_data: MemberCreate):
        """Create member profile (admin only).

        Supports 2 paths:
        1) Legacy: provide user_id + member_code
        2) Admin onboarding: provide full_name/contact and create or link user automatically
        """
        normalized_code = MemberService._normalize_member_code(
            member_data.member_code or member_data.member_id
        )

        # Legacy path: explicit user_id linking
        if member_data.user_id is not None:
            existing_user = db.query(User).filter(User.id == member_data.user_id).first()
            if not existing_user:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="User not found",
                )

            existing_member = db.query(Member).filter(Member.user_id == member_data.user_id).first()
            if existing_member:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Member profile already exists for this user",
                )

            member_code = normalized_code or MemberService._next_member_code(db)
            duplicate_code = db.query(Member).filter(Member.member_code == member_code).first()
            if duplicate_code:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Member code already in use",
                )

            member = Member(
                user_id=member_data.user_id,
                member_code=member_code,
                monthly_amount=member_data.monthly_amount,
                address=member_data.address,
                status=member_data.status or "active",
                join_date=member_data.join_date,
            )
            db.add(member)
            db.commit()
            db.refresh(member)
            return member

        # Admin onboarding path for offline members.
        full_name = MemberService._normalize_contact(member_data.full_name)
        phone = MemberService._normalize_contact(member_data.phone)
        email = MemberService._normalize_contact(member_data.email)

        if not full_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="full_name is required for admin member onboarding",
            )

        if not phone and not email:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least one contact detail (phone or email) is required",
            )

        member_code = normalized_code or MemberService._next_member_code(db)
        duplicate_code = db.query(Member).filter(Member.member_code == member_code).first()
        if duplicate_code:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Member code already in use",
            )

        user, _created_user = MemberService._get_or_create_user_for_admin_member(
            db=db,
            full_name=full_name,
            phone=phone,
            email=email,
            member_code=member_code,
        )

        existing_member = db.query(Member).filter(Member.user_id == user.id).first()
        if existing_member:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A member profile already exists for this contact/user",
            )

        # Keep imported/offline users non-login by default until they claim via invite registration.
        if user.role != "member":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Cannot link member profile to a non-member user role",
            )

        if user.username.startswith("offline_") or not user.is_active:
            user.username = MemberService._generate_unique_offline_username(db, full_name, member_code)
            user.is_active = False

        if email and not user.email:
            user.email = email
        if phone and not user.phone:
            user.phone = phone

        member = Member(
            user_id=user.id,
            member_code=member_code,
            monthly_amount=member_data.monthly_amount,
            address=member_data.address,
            status=member_data.status or "active",
            join_date=member_data.join_date,
        )

        db.add(member)
        db.commit()
        db.refresh(member)
        return member

    @staticmethod
    def update_member(db: Session, member_id: int, update_data: MemberUpdate):
        """Update member information and linked user contact fields."""
        member = MemberService.get_member(db, member_id)
        user = db.query(User).filter(User.id == member.user_id).first()

        update_fields = update_data.dict(exclude_unset=True)

        new_code = MemberService._normalize_member_code(update_fields.get("member_code") or update_fields.get("member_id"))
        if new_code and new_code != member.member_code:
            duplicate_code = db.query(Member).filter(Member.member_code == new_code, Member.id != member.id).first()
            if duplicate_code:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Member code already in use",
                )
            member.member_code = new_code

        if "full_name" in update_fields and update_fields["full_name"] and user:
            user.username = update_fields["full_name"].strip()

        if "email" in update_fields and user:
            new_email = MemberService._normalize_contact(update_fields.get("email"))
            if new_email and new_email != user.email:
                duplicate_email = db.query(User).filter(User.email == new_email, User.id != user.id).first()
                if duplicate_email:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="Email already in use",
                    )
            user.email = new_email

        if "phone" in update_fields and user:
            new_phone = MemberService._normalize_contact(update_fields.get("phone"))
            if new_phone and new_phone != user.phone:
                duplicate_phone = db.query(User).filter(User.phone == new_phone, User.id != user.id).first()
                if duplicate_phone:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="Phone already in use",
                    )
            user.phone = new_phone

        for key in ("monthly_amount", "address", "status", "join_date"):
            if key in update_fields and update_fields[key] is not None:
                setattr(member, key, update_fields[key])

        db.commit()
        db.refresh(member)

        return member

    @staticmethod
    def delete_member(db: Session, member_id: int):
        """Delete a member profile and all related records (superadmin only)."""
        member = MemberService.get_member(db, member_id)
        # Delete related challans and bulk groups first to satisfy FK constraints
        db.query(Challan).filter(Challan.member_id == member.id).delete(synchronize_session=False)
        db.query(BulkChallanGroup).filter(BulkChallanGroup.member_id == member.id).delete(synchronize_session=False)
        db.flush()
        db.delete(member)
        db.commit()
        return None

    @staticmethod
    def get_member_for_user(db: Session, user_id: int):
        """Get member associated with user."""
        member = db.query(Member).filter(Member.user_id == user_id).first()

        if not member:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Member profile not found for this user",
            )

        return member

    @staticmethod
    def _read_tabular_rows(file_bytes: bytes, filename: str) -> list[dict]:
        lower_name = filename.lower()

        if lower_name.endswith(".csv"):
            text = file_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return [dict(row) for row in reader]

        if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
            try:
                from openpyxl import load_workbook  # type: ignore[import-not-found]
            except ImportError as exc:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Excel import requires openpyxl package",
                ) from exc

            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            records: list[dict] = []
            for raw in rows[1:]:
                if raw is None:
                    continue
                record = {}
                empty_row = True
                for idx, value in enumerate(raw):
                    key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
                    if value not in (None, ""):
                        empty_row = False
                    record[key] = value
                if not empty_row:
                    records.append(record)
            return records

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use .csv or .xlsx",
        )

    @staticmethod
    def _normalized_row(row: dict) -> dict:
        normalized = {}
        for key, value in row.items():
            if key is None:
                continue
            nk = str(key).strip().lower().replace(" ", "_")
            normalized[nk] = value
        return normalized

    @staticmethod
    def _row_value(row: dict, keys: list[str]):
        for key in keys:
            if key in row and row[key] not in (None, ""):
                return row[key]
        return None

    @staticmethod
    def import_members_file(
        db: Session,
        file_bytes: bytes,
        filename: str,
        include_donations: bool = True,
        imported_by_user_id: Optional[int] = None,
    ) -> MemberImportSummary:
        rows = MemberService._read_tabular_rows(file_bytes, filename)

        total_rows = len(rows)
        members_created = 0
        members_linked_existing = 0
        challans_created = 0
        rows_skipped = 0
        errors: list[str] = []

        for idx, raw_row in enumerate(rows, start=2):
            row = MemberService._normalized_row(raw_row)

            try:
                username = MemberService._normalize_contact(
                    MemberService._row_value(row, ["username", "user_name"])
                )
                si_code = MemberService._si_to_member_code(
                    MemberService._row_value(row, ["si_no", "si", "serial_no"])
                )

                member_code = MemberService._normalize_member_code(
                    MemberService._row_value(row, ["member_code", "member_id", "code"])
                ) or si_code
                full_name = MemberService._normalize_contact(
                    MemberService._row_value(row, ["full_name", "name", "member_name"])
                )
                phone = MemberService._normalize_contact(
                    MemberService._row_value(row, ["phone", "mobile", "phone_number"])
                )
                email = MemberService._normalize_contact(
                    MemberService._row_value(row, ["email", "email_address"])
                )
                address = MemberService._normalize_contact(
                    MemberService._row_value(row, ["address", "location", "notes"])
                )
                monthly_amount_raw = MemberService._row_value(row, ["monthly_amount", "monthly", "default_amount"])
                status_value = MemberService._normalize_contact(MemberService._row_value(row, ["status"])) or "active"
                join_date = MemberService._parse_datetime(
                    MemberService._row_value(row, ["join_date", "joined_on", "member_since"])
                )
                if not join_date:
                    join_year = MemberService._row_value(row, ["join_year", "year"])
                    if join_year not in (None, ""):
                        try:
                            join_date = datetime(int(float(str(join_year).strip())), 1, 1)
                        except (TypeError, ValueError):
                            join_date = None

                monthly_amount = 0.0
                if monthly_amount_raw not in (None, ""):
                    monthly_amount = float(monthly_amount_raw)

                has_member_profile_data = bool(full_name or phone or email or monthly_amount_raw not in (None, ""))

                if not has_member_profile_data and not username and not member_code:
                    rows_skipped += 1
                    errors.append(f"Row {idx}: missing identifier fields")
                    continue

                member = MemberService._find_member_by_identifiers(
                    db=db,
                    member_code=member_code,
                    username=username,
                    phone=phone,
                    email=email,
                )

                if member:
                    members_linked_existing += 1
                elif has_member_profile_data:
                    if not member_code:
                        member_code = MemberService._next_member_code(db)

                    user, created_new_user = MemberService._get_or_create_user_for_admin_member(
                        db=db,
                        full_name=full_name or username or "Member",
                        phone=phone,
                        email=email,
                        member_code=member_code,
                    )

                    if username and user.username.startswith("offline_"):
                        duplicate_username = db.query(User).filter(User.username == username, User.id != user.id).first()
                        if not duplicate_username:
                            user.username = username

                    member = db.query(Member).filter(Member.user_id == user.id).first()
                    if member:
                        members_linked_existing += 1
                    else:
                        member = Member(
                            user_id=user.id,
                            member_code=member_code,
                            monthly_amount=monthly_amount,
                            address=address,
                            status=status_value,
                            join_date=join_date,
                        )
                        db.add(member)
                        db.flush()
                        members_created += 1
                        if not created_new_user:
                            members_linked_existing += 1

                if not member:
                    rows_skipped += 1
                    errors.append(f"Row {idx}: member not found and insufficient profile data to create one")
                    continue

                if include_donations:
                    donation_type = MemberService._normalize_contact(
                        MemberService._row_value(row, ["type", "donation_type"])
                    )
                    donation_month = MemberService._normalize_contact(
                        MemberService._row_value(row, ["month", "donation_month", "payment_month", "period"])
                    )
                    donation_month = MemberService._parse_month(donation_month)
                    donation_amount_raw = MemberService._row_value(row, ["amount", "donation_amount", "paid_amount"])
                    payment_method = MemberService._normalize_contact(
                        MemberService._row_value(row, ["payment_method", "method"])
                    )
                    donation_status_raw = MemberService._normalize_contact(
                        MemberService._row_value(row, ["donation_status", "status", "challan_status", "payment_status"])
                    )

                    if donation_amount_raw not in (None, ""):
                        donation_amount = float(donation_amount_raw)

                        normalized_status = "generated"
                        raw_status = (donation_status_raw or "generated").lower()
                        if raw_status in ("approved", "paid", "completed"):
                            normalized_status = "approved"
                        elif raw_status in ("pending",):
                            normalized_status = "pending"
                        elif raw_status in ("rejected", "failed"):
                            normalized_status = "rejected"

                        challan_type = "campaign" if (donation_type or "").lower() == "campaign" else "monthly"

                        if challan_type == "monthly" and not donation_month:
                            rows_skipped += 1
                            errors.append(f"Row {idx}: missing valid month for monthly donation")
                            continue

                        campaign_id = None
                        if challan_type == "campaign":
                            campaign_id = MemberService._resolve_campaign_for_import(
                                db,
                                MemberService._row_value(row, ["suggested_campaign_name", "campaign_name"]),
                                imported_by_user_id,
                            )

                        duplicate_query = db.query(Challan).filter(
                            Challan.member_id == member.id,
                            Challan.type == challan_type,
                            Challan.amount == donation_amount,
                        )
                        if challan_type == "monthly":
                            duplicate_query = duplicate_query.filter(Challan.month == donation_month)
                        else:
                            duplicate_query = duplicate_query.filter(Challan.campaign_id == campaign_id)
                        duplicate_challan = duplicate_query.first()

                        if not duplicate_challan:
                            challan = Challan(
                                member_id=member.id,
                                type=challan_type,
                                month=donation_month if challan_type == "monthly" else None,
                                campaign_id=campaign_id,
                                amount=donation_amount,
                                payment_method=payment_method,
                                status=normalized_status,
                            )
                            if normalized_status == "approved":
                                challan.approved_at = datetime.utcnow()
                            db.add(challan)
                            challans_created += 1

            except (ValueError, TypeError) as exc:
                rows_skipped += 1
                errors.append(f"Row {idx}: {exc}")

        db.commit()

        return MemberImportSummary(
            total_rows=total_rows,
            members_created=members_created,
            members_linked_existing=members_linked_existing,
            challans_created=challans_created,
            rows_skipped=rows_skipped,
            errors=errors[:50],
        )